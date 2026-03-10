import os
import re
import sys
import glob
import hashlib
import getpass
import json
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


VERSION = "v3.2"
AUTH_CONFIG_FILE = "auth_config.json"
# 固定盐（pepper）：与密码拼接后再哈希，避免仅靠 json 哈希被直接撞库
AUTH_PEPPER = "MHG@2026#scan-query#pepper"


def clear():
    os.system("cls" if os.name == "nt" else "clear")


def sha256_hex(text: str) -> str:
    return hashlib.sha256(str(text or "").encode("utf-8")).hexdigest()


def hash_password(password: str) -> str:
    mixed = f"{AUTH_PEPPER}{password}{AUTH_PEPPER}"
    return sha256_hex(mixed)


def load_auth_hash(config_path: str) -> str:
    if not os.path.isfile(config_path):
        return ""
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return str(data.get("password_sha256", "")).strip().lower()
    except Exception:
        return ""


def check_password_or_exit(max_try: int = 3):
    """启动密码校验：读取同目录 auth_config.json 的 password_sha256。"""
    current_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    config_path = os.path.join(current_dir, AUTH_CONFIG_FILE)
    target_hash = load_auth_hash(config_path)

    if not target_hash:
        print(f"\n[WARN] 未读取到密码配置: {config_path}")
        print("  请先运行 set_password.py 设置密码后再使用。")
        raise SystemExit(1)

    print("\n[AUTH] 请输入启动密码")
    for i in range(1, max_try + 1):
        try:
            pwd = getpass.getpass("密码: ")
        except Exception:
            pwd = input("密码: ").strip()

        if hash_password(pwd).lower() == target_hash:
            print("[OK] 密码正确，开始运行。\n")
            return

        left = max_try - i
        if left > 0:
            print(f"[ERR] 密码错误，还可尝试 {left} 次")
            time.sleep(1.2)

    print("[ERR] 密码连续错误，程序已退出。")
    raise SystemExit(1)


def print_header():
    print("=" * 60)
    print(f"  棉花果订单扫描工具  (EXE版)  {VERSION}")
    print("  生成 Excel 供后续查询使用")
    print("=" * 60)
    print("[STEP 0/3] 初始化")
    print()


def ask(prompt, default=""):
    val = input(prompt).strip()
    return val if val else default


def read_cell(ws, addr):
    val = ws[addr].value
    if val is None:
        return ""
    return str(val).strip()


def resolve_username(ws):
    """
    用户名读取规则：
    - 常规：F5 即用户名
    - 兼容：若 F5 是“客户名/用户名”这类标签，则读取 G5
    """
    f5 = read_cell(ws, "F5")
    g5 = read_cell(ws, "G5")

    label_words = {"客户名", "用户名", "客户", "用户", "姓名"}
    if f5 in label_words and g5:
        return g5

    return f5 or g5


def read_order_info(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        username = resolve_username(ws)
        order_id = read_cell(ws, "J2")
        wb.close()
        return username, order_id
    except Exception:
        return None, None


def parse_order_base_id(order_id: str):
    """
    提取订单基础编号（用于区间比较）
    例如:
      99481   -> 99481
      99550-1 -> 99550
      99556-1 -> 99556
    无法提取则返回 None
    """
    text = str(order_id or "").strip()
    if not text:
        return None

    m = re.match(r"^(\d+)", text)
    if not m:
        return None

    try:
        return int(m.group(1))
    except Exception:
        return None


def normalize_digits(text: str) -> str:
    return re.sub(r"\D", "", str(text or ""))


def is_buou_pattern(raw_name: str, name_digits: str) -> bool:
    """
    布偶人识别策略（按用户名本身判断，不依赖订单号）：
    - 用户名去掉符号后是纯数字
    - 数字长度 >= 8（如 2606060361）
    - 原始用户名不包含中文/英文（避免误判普通姓名）

    说明：你当前业务里“数字型用户名”基本就是该客户，
    用这个策略比“跟订单号相似”更稳。
    """
    if len(name_digits) < 8:
        return False

    # 包含中文或英文字母则不判定为该特殊客户
    if re.search(r"[A-Za-z\u4e00-\u9fff]", raw_name):
        return False

    # 允许数字、空格和常见分隔符（-_/）
    if not re.fullmatch(r"[\d\s\-_/]+", raw_name):
        return False

    return True


def mark_special_customer(username: str, _order_id: str) -> str:
    """
    若识别为“布偶人”的数字型编号用户名，则自动前缀“布偶人”。
    """
    raw_name = str(username or "").strip()
    if not raw_name:
        return raw_name

    # 已经加过前缀则不重复
    if raw_name.startswith("布偶人"):
        return raw_name

    # 仅对“数字型用户名”生效
    name_digits = normalize_digits(raw_name)
    if not name_digits:
        return raw_name

    if not is_buou_pattern(raw_name, name_digits):
        return raw_name

    return f"布偶人{raw_name}"


def list_excel_files(folder_path):
    """扫描该编号文件夹下所有 Excel 文件（xlsm 优先，其次 xlsx）。"""
    all_files = []
    all_files.extend(glob.glob(os.path.join(folder_path, "*.xlsm")))
    all_files.extend(glob.glob(os.path.join(folder_path, "*.xlsx")))

    def rank(path):
        name = os.path.basename(path)
        ext = os.path.splitext(name)[1].lower()
        ext_rank = 0 if ext == ".xlsm" else 1
        return (ext_rank, name)

    all_files.sort(key=rank)
    return all_files


def pick_best_order_info(file_list):
    """
    在所有 Excel 中逐个尝试：
    - 先看 J2 是否为“有效编号”（可解析出数字前缀）
    - 有效后再取用户名（F5/G5 规则不变）
    - 优先返回“编号+用户名”都拿到的结果
    - 否则返回“至少有有效编号”的结果
    """
    best_username = ""
    best_order_id = ""
    best_file = ""

    for fp in file_list:
        username, order_id = read_order_info(fp)
        username = (username or "").strip()
        order_id = (order_id or "").strip()

        # 关键：J2 必须可识别为编号（例如 99556 / 99556-1）
        if parse_order_base_id(order_id) is None:
            continue

        if order_id and username:
            return username, order_id, fp

        if order_id and not best_order_id:
            best_order_id = order_id
            best_username = username
            best_file = fp

    if best_order_id:
        return best_username, best_order_id, best_file

    return "", "", ""


def scan_orders(root_path, id_min, id_max):
    t0 = time.time()

    if not os.path.isdir(root_path):
        print(f"[错误] 路径不存在: {root_path}")
        return []

    subdirs_all = sorted([
        d for d in os.listdir(root_path)
        if os.path.isdir(os.path.join(root_path, d))
    ])

    # 关键优化：若一级目录名本身就是编号，则先按区间过滤目录，避免全量扫描
    if id_min is not None or id_max is not None:
        subdirs = []
        skipped_by_range = 0
        for d in subdirs_all:
            base_from_dir = parse_order_base_id(d)
            if base_from_dir is None:
                # 无法从目录名提取编号：保守起见仍纳入扫描
                subdirs.append(d)
                continue

            if id_min is not None and base_from_dir < id_min:
                skipped_by_range += 1
                continue
            if id_max is not None and base_from_dir > id_max:
                skipped_by_range += 1
                continue

            subdirs.append(d)

        print(f"\n发现 {len(subdirs_all)} 个子文件夹，按区间预筛后需扫描 {len(subdirs)} 个（已跳过 {skipped_by_range} 个）...\n")
    else:
        subdirs = subdirs_all
        print(f"\n发现 {len(subdirs)} 个子文件夹，开始扫描...\n")

    results = []
    skipped = 0

    for i, folder_name in enumerate(subdirs, 1):
        folder_path = os.path.join(root_path, folder_name)

        excel_files = list_excel_files(folder_path)
        if not excel_files:
            skipped += 1
            if i % 20 == 0:
                print(f"  进度 [{i}/{len(subdirs)}] ...")
            continue

        username, order_id, matched_file = pick_best_order_info(excel_files)
        if not order_id:
            skipped += 1
            if i % 20 == 0:
                print(f"  进度 [{i}/{len(subdirs)}] ...")
            continue

        # 特定客户识别：数字型用户名 + 编号特征 -> 自动加“布偶人”前缀
        username = mark_special_customer(username, order_id)

        # 区间过滤：按“基础编号”比较，兼容 99550-1 这种格式
        base_id = parse_order_base_id(order_id)
        if base_id is not None:
            if id_min is not None and base_id < id_min:
                continue
            if id_max is not None and base_id > id_max:
                continue
        elif id_min is not None or id_max is not None:
            # 用户设置了区间，但该编号无法解析为数字，直接跳过，避免混入无关数据
            continue

        results.append({
            "用户名": username,
            "订单编号": order_id,
            "基础编号": base_id if base_id is not None else "",
            "状态": "已找到",
        })

        file_hint = os.path.basename(matched_file) if matched_file else "(候选文件中解析)"
        print(
            f"  [{i}/{len(subdirs)}] ✓ 编号:{order_id}  基础编号:{base_id if base_id is not None else '-'}  "
            f"用户:{username or '(无)'}  文件:{file_hint}"
        )

    elapsed = time.time() - t0
    total_scanned = len(subdirs)
    avg = (elapsed / total_scanned) if total_scanned > 0 else 0.0

    print(f"\n扫描完毕：找到 {len(results)} 条，跳过 {skipped} 个无有效Excel/无有效编号文件夹")
    print(f"耗时统计：总耗时 {elapsed:.1f} 秒，平均每目录 {avg:.3f} 秒")

    return results


def fill_missing_ids(order_list, id_min, id_max):
    """若给了起止编号，补齐区间内缺失编号并打标记。"""
    if id_min is None or id_max is None:
        return sorted(
            order_list,
            key=lambda x: (
                x.get("基础编号") if isinstance(x.get("基础编号"), int) else 10**12,
                str(x.get("订单编号", "")),
            ),
        )

    if id_min > id_max:
        id_min, id_max = id_max, id_min

    present_base_ids = {
        item.get("基础编号")
        for item in order_list
        if isinstance(item.get("基础编号"), int)
    }

    completed = list(order_list)
    for n in range(id_min, id_max + 1):
        if n not in present_base_ids:
            completed.append({
                "用户名": "",
                "订单编号": str(n),
                "基础编号": n,
                "状态": "缺失(未找到文件)",
            })

    completed.sort(
        key=lambda x: (
            x.get("基础编号") if isinstance(x.get("基础编号"), int) else 10**12,
            str(x.get("订单编号", "")),
        )
    )
    return completed


def save_excel(order_list, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单列表"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    fill_found = PatternFill("solid", start_color="E2EFDA")
    fill_missing = PatternFill("solid", start_color="FCE4D6")

    # 表头
    headers = ["用户名", "订单编号(oId)", "基础编号", "状态"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="4472C4")
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 22

    # 数据
    for row_idx, item in enumerate(order_list, 2):
        status = item.get("状态", "")
        row_fill = fill_missing if "缺失" in status else fill_found

        row_values = [
            item.get("用户名", ""),
            item.get("订单编号", ""),
            item.get("基础编号", ""),
            status,
        ]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = center
            cell.border = border
            cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 18

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.freeze_panes = "A2"

    wb.save(output_path)


def main():
    clear()
    print_header()
    check_password_or_exit()

    # 输入路径
    current_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    root_path = ask(f"网盘路径（直接回车使用当前目录）\n> ", current_dir)

    if not os.path.isdir(root_path):
        print(f"\n[错误] 路径不存在: {root_path}")
        return

    print()

    # 输入区间
    min_str = ask("起始编号（直接回车不限制）: ")
    max_str = ask("结束编号（直接回车不限制）: ")

    id_min = int(min_str) if min_str.isdigit() else None
    id_max = int(max_str) if max_str.isdigit() else None

    range_desc = "全部"
    if id_min is not None and id_max is not None:
        range_desc = f"{id_min} ~ {id_max}"
    elif id_min is not None:
        range_desc = f">= {id_min}"
    elif id_max is not None:
        range_desc = f"<= {id_max}"

    print(f"\n[STEP 1/3] 参数确认")
    print(f"路径: {root_path}")
    print(f"区间: {range_desc}")
    input("\n按回车开始扫描...")
    print()

    # 扫描
    print("[STEP 2/3] 扫描订单")
    order_list = scan_orders(root_path, id_min, id_max)

    # 补齐缺失编号
    order_list = fill_missing_ids(order_list, id_min, id_max)

    if not order_list:
        print("\n没有找到任何符合条件的订单。")
        return

    # 保存
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(current_dir, f"orders_{now}.xlsx")
    save_excel(order_list, output_path)

    total_missing = sum(1 for x in order_list if "缺失" in str(x.get("状态", "")))
    print("\n[STEP 3/3] 输出结果")
    print(f"[OK] 已保存 {len(order_list)} 条记录到:")
    print(f"   {output_path}")
    print(f"   其中缺失编号: {total_missing} 条")

    # 自动打开文件夹
    try:
        os.startfile(os.path.dirname(output_path))
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[ERR] 运行异常: {e}")
    finally:
        input("\n按回车退出...")
