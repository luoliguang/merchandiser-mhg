"""
棉花果订单一站式查询工具 v3.0
=====================================
功能：
  1. 扫描指定网盘路径下所有子文件夹
  2. 从每个文件夹的 xlsx/xlsm 订单表中读取用户名(F5) 和订单编号(J2)
  3. 支持按编号区间过滤（如只处理 40000~40050）
  4. 自动去棉花果查询进度，输出彩色结果 Excel

运行方式: python mhg_full.py
"""

import os
import glob
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time


# ============================================================
#  ★ 配置区 —— 每次只需改这里 ★
# ============================================================
CONFIG = {
    # 网盘根目录路径（包含所有订单子文件夹的那一层）
    # 例如: "D:\\网盘\\订单" 或 "\\\\192.168.1.10\\share\\订单"
    "root_path": r"D:\Data\Data\GitHub\mianhuaguo\模拟网盘",

    # 编号区间过滤（填数字）：只处理这个范围内的编号
    # 如果不想限制范围，把两个都设为 None
    "order_id_min": 99430,   # 起始编号（含）
    "order_id_max": 99432,   # 结束编号（含）

    # 棉花果 Cookie（过期后重新从 F12 复制）
    "cookie": "PHPSESSID=39js5qvoc98b99pov580eu9a5i",

    # 查询日期范围
    "start_date": "2024-01-01",
    "end_date":   "2026-12-31",

    # 每次请求间隔（秒），避免请求过快
    "request_delay": 1.0,

    # 输出文件路径（留空则保存在当前目录）
    "output_excel": f"orders_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
}
# ============================================================


API_GET_LIST  = "https://58mhg.com/index/orders/getList"
API_GET_COUNT = "https://58mhg.com/index/orders/getOrdersCount"

COLOR_NOT_FOUND = "FCE4D6"  # 红：查无此单
COLOR_PENDING   = "DDEBF7"  # 蓝：待开始
COLOR_IN_PROG   = "FFF2CC"  # 黄：进行中
COLOR_DONE      = "E2EFDA"  # 绿：已完成
COLOR_READ_ERR  = "EDEDED"  # 灰：文件读取失败


# ─────────────────────────────────────────
#  1. 扫描网盘，读取订单编号和用户名
# ─────────────────────────────────────────

def read_cell(ws, cell_addr: str) -> str:
    """读取单元格值，返回字符串，空则返回空字符串"""
    val = ws[cell_addr].value
    if val is None:
        return ""
    return str(val).strip()


def read_order_info_from_file(filepath: str):
    """
    从 xlsx 或 xlsm 文件中读取：
      F5 → 用户名
      J2 → 订单编号
    返回 (username, order_id) 或 (None, None) 表示读取失败
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        username = read_cell(ws, "F5")
        order_id = read_cell(ws, "J2")
        wb.close()
        return username, order_id
    except Exception as e:
        return None, None


def scan_orders(root_path: str, id_min, id_max) -> list:
    """
    扫描 root_path 下所有子文件夹，读取订单信息。
    优先读 xlsx，xlsx 的 J2 为空则再读 xlsm。
    按编号区间过滤。
    返回 [{"用户名": ..., "订单编号": ..., "来源文件": ...}, ...]
    """
    if not os.path.isdir(root_path):
        raise RuntimeError(f"路径不存在或无法访问: {root_path}")

    results = []
    # 获取所有直接子文件夹
    subdirs = [
        d for d in os.listdir(root_path)
        if os.path.isdir(os.path.join(root_path, d))
    ]
    subdirs.sort()

    print(f"📁 共发现 {len(subdirs)} 个子文件夹")

    for folder_name in subdirs:
        folder_path = os.path.join(root_path, folder_name)

        # 查找 xlsx 和 xlsm 文件
        xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
        xlsm_files = glob.glob(os.path.join(folder_path, "*.xlsm"))

        username, order_id, source = "", "", ""

        # 优先读 xlsx
        if xlsx_files:
            u, oid = read_order_info_from_file(xlsx_files[0])
            if oid:
                username, order_id, source = u or "", oid, xlsx_files[0]

        # xlsx J2 为空时，尝试 xlsm
        if not order_id and xlsm_files:
            u, oid = read_order_info_from_file(xlsm_files[0])
            if oid:
                username, order_id, source = u or "", oid, xlsm_files[0]

        if not order_id:
            print(f"  ⚠ [{folder_name}] 未读取到订单编号，跳过")
            continue

        # 编号区间过滤
        try:
            oid_num = int(order_id)
            if id_min is not None and oid_num < id_min:
                continue
            if id_max is not None and oid_num > id_max:
                continue
        except ValueError:
            pass  # 非纯数字编号不过滤，直接保留

        results.append({
            "用户名":   username,
            "订单编号": order_id,
            "来源文件": source,
        })

    return results


# ─────────────────────────────────────────
#  2. 棉花果查询
# ─────────────────────────────────────────

def build_headers():
    return {
        "Content-Type": "application/json",
        "Cookie": CONFIG["cookie"],
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://58mhg.com/console/",
        "Origin": "https://58mhg.com",
    }


def build_payload(keyword: str, page_index: int = 1, page_size: int = 20):
    return {
        "pageIndex": page_index,
        "pageSize": page_size,
        "pageItem": 1,
        "keyWord": str(keyword),
        "showMaterial": True,
        "withDepartment": False,
        "startDate": CONFIG["start_date"],
        "endDate": CONFIG["end_date"],
        "startOut": "",
        "endOut": "",
        "key": "like",
        "isBuild": True,
    }


def process_to_status(process: float):
    if process >= 1.0:
        return "已完成", COLOR_DONE
    elif process == 0.0:
        return "待开始", COLOR_PENDING
    elif process == 0.25:
        return "生产中 25%", COLOR_IN_PROG
    elif process == 0.5:
        return "生产中 50%", COLOR_IN_PROG
    elif process == 0.75:
        return "生产中 75%", COLOR_IN_PROG
    else:
        return f"生产中 {int(process*100)}%", COLOR_IN_PROG


def query_order(keyword: str):
    """带重试的查询，返回 (found: bool, orders: list)"""
    MAX_RETRY = 3
    RETRY_WAIT = 3

    for attempt in range(1, MAX_RETRY + 1):
        try:
            session = requests.Session()
            headers = build_headers()

            count_resp = session.post(
                API_GET_COUNT,
                json=build_payload(keyword),
                headers=headers,
                timeout=15,
            )
            count_data = count_resp.json()

            if count_data.get("code") != 200:
                print(f"    ⚠ Cookie可能已过期，请重新复制")
                return False, []

            total = int(count_data.get("data", 0))
            if total == 0:
                return False, []

            page_size = 20
            pages = max(1, -(-total // page_size))
            results = []

            for page in range(1, pages + 1):
                resp = session.post(
                    API_GET_LIST,
                    json=build_payload(keyword, page_index=page, page_size=page_size),
                    headers=headers,
                    timeout=15,
                )
                data = resp.json()
                if data.get("code") != 200:
                    break
                orders = data.get("data", [])
                results.extend(orders)
                if len(orders) < page_size:
                    break

            return True, results

        except requests.exceptions.ConnectionError:
            print(f"    ⚠ 连接失败 ({attempt}/{MAX_RETRY})，{RETRY_WAIT}秒后重试...")
        except requests.exceptions.Timeout:
            print(f"    ⚠ 请求超时 ({attempt}/{MAX_RETRY})，{RETRY_WAIT}秒后重试...")
        except Exception as e:
            print(f"    ⚠ 出错 ({attempt}/{MAX_RETRY}): {e}")

        if attempt < MAX_RETRY:
            time.sleep(RETRY_WAIT)

    print(f"    ❌ 重试{MAX_RETRY}次均失败")
    return False, []


# ─────────────────────────────────────────
#  3. 保存结果 Excel
# ─────────────────────────────────────────

def save_results(all_rows: list, filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "查询结果"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="4472C4")
    normal_font = Font(name="Arial", size=10)

    headers = ["用户名", "查询编号", "订单内部ID", "产品名称",
               "颜色/规格", "总数量", "订单状态", "进度百分比", "下单时间", "查询时间"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(all_rows, 2):
        row_fill = PatternFill("solid", start_color=row.get("_color", COLOR_NOT_FOUND))
        values = [
            row.get("用户名", ""),
            row.get("查询编号", ""),
            row.get("oId", ""),
            row.get("name", ""),
            row.get("color_size", ""),
            row.get("totalQuantity", ""),
            row.get("状态文字", ""),
            row.get("进度百分比", ""),
            row.get("orderTime", ""),
            row.get("查询时间", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = center
            cell.border = border
            cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 18

    col_widths = [12, 14, 14, 24, 18, 10, 14, 12, 18, 18]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # 图例说明
    ws2 = wb.create_sheet("颜色说明")
    legend = [
        (COLOR_DONE,      "✅ 绿色", "已完成"),
        (COLOR_IN_PROG,   "🟡 黄色", "生产进行中"),
        (COLOR_PENDING,   "🔵 蓝色", "待开始（订单存在，尚未生产）"),
        (COLOR_NOT_FOUND, "🔴 红色", "查无此单"),
        (COLOR_READ_ERR,  "⬜ 灰色", "文件读取失败"),
    ]
    for i, (color, label, desc) in enumerate(legend, 1):
        fill = PatternFill("solid", start_color=color)
        for col_idx, val in enumerate([label, desc], 1):
            c = ws2.cell(row=i, column=col_idx, value=val)
            c.fill = fill
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[i].height = 22
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 38

    wb.save(filepath)
    print(f"✅ 结果已保存: {filepath}")


# ─────────────────────────────────────────
#  4. 主流程
# ─────────────────────────────────────────

def main():
    print("=" * 60)
    print("  棉花果订单一站式查询工具  v3.0")
    print("=" * 60)

    # 校验配置
    if not os.path.isdir(CONFIG["root_path"]):
        print(f"\n❌ 网盘路径不存在: {CONFIG['root_path']}")
        print("   请修改 CONFIG 中的 root_path")
        return

    if not CONFIG["cookie"] or len(CONFIG["cookie"]) < 10:
        print("\n❌ 请先在 CONFIG 中填入你的 Cookie")
        return

    id_min = CONFIG.get("order_id_min")
    id_max = CONFIG.get("order_id_max")

    range_desc = "全部编号"
    if id_min is not None and id_max is not None:
        range_desc = f"编号区间 {id_min} ~ {id_max}"
    elif id_min is not None:
        range_desc = f"编号 ≥ {id_min}"
    elif id_max is not None:
        range_desc = f"编号 ≤ {id_max}"

    print(f"\n📁 网盘路径: {CONFIG['root_path']}")
    print(f"🔢 查询范围: {range_desc}")
    print(f"\n{'─'*60}")

    # Step 1: 扫描网盘
    print("\n【第一步】扫描网盘，读取订单信息...")
    try:
        order_list = scan_orders(CONFIG["root_path"], id_min, id_max)
    except RuntimeError as e:
        print(f"❌ {e}")
        return

    if not order_list:
        print("⚠ 在指定范围内没有找到任何订单，请检查路径和区间设置")
        return

    print(f"✅ 符合条件的订单: {len(order_list)} 条\n")

    # Step 2: 逐一查询棉花果
    print("【第二步】查询棉花果订单进度...")
    print(f"{'─'*60}")

    all_rows = []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cnt_done = cnt_pending = cnt_inprog = cnt_notfound = 0
    total = len(order_list)

    for idx, item in enumerate(order_list):
        username = item["用户名"]
        keyword  = item["订单编号"]

        print(f"[{idx+1}/{total}] 编号: {keyword}  用户: {username or '(无)'}")

        found, orders = query_order(keyword)

        if not found:
            cnt_notfound += 1
            print(f"    → ❌ 查无此单")
            all_rows.append({
                "用户名": username, "查询编号": keyword,
                "oId": "", "name": "查无此单",
                "color_size": "", "totalQuantity": "",
                "状态文字": "查无此单", "进度百分比": "",
                "orderTime": "", "查询时间": now_str,
                "_color": COLOR_NOT_FOUND,
            })
        else:
            for order in orders:
                raw_process = order.get("process")
                process = float(raw_process) if raw_process is not None else 0.0
                status_text, row_color = process_to_status(process)

                if process >= 1.0:
                    cnt_done += 1
                elif process == 0.0:
                    cnt_pending += 1
                else:
                    cnt_inprog += 1

                quantities = order.get("quantities", {})
                color_size = "  ".join(
                    f"{c}/{s}×{q}"
                    for c, sizes in quantities.items()
                    for s, q in sizes.items()
                )

                order_ts = order.get("orderTime")
                order_time = ""
                if order_ts:
                    try:
                        order_time = datetime.fromtimestamp(order_ts / 1000).strftime("%Y-%m-%d %H:%M")
                    except:
                        pass

                all_rows.append({
                    "用户名": username,
                    "查询编号": keyword,
                    "oId": order.get("oId", ""),
                    "name": order.get("name", ""),
                    "color_size": color_size,
                    "totalQuantity": order.get("totalQuantity", ""),
                    "状态文字": status_text,
                    "进度百分比": f"{int(process*100)}%",
                    "orderTime": order_time,
                    "查询时间": now_str,
                    "_color": row_color,
                })
                print(f"    → {order.get('name','')} | {status_text} | 数量: {order.get('totalQuantity','')}")

        time.sleep(CONFIG["request_delay"])

    # Step 3: 保存
    print(f"\n{'─'*60}")
    print("【第三步】保存结果...")
    save_results(all_rows, CONFIG["output_excel"])

    # 摘要
    print(f"\n{'='*60}")
    print(f"  全部完成！处理 {total} 个编号，共 {len(all_rows)} 条记录")
    print(f"  ✅ 已完成:   {cnt_done} 条")
    print(f"  🔵 待开始:   {cnt_pending} 条")
    print(f"  🟡 进行中:   {cnt_inprog} 条")
    print(f"  ❌ 查无此单: {cnt_notfound} 条")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
